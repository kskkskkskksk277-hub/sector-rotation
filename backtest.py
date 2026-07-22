# -*- coding: utf-8 -*-
"""
============================================================
 シグナルのバックテスト（イベントスタディ）
============================================================
 ダッシュボードの6種シグナル（ゼロクロス↑↓・±σブレイク・加速/減速）の
 発生後に各資産を買った場合の 1ヶ月/3ヶ月/6ヶ月 先のリターンを検証し、
 Excel（サマリー＋取引一覧）に出力します。

 検証対象の資産:
   ・日経平均 / TOPIX
   ・攻めバスケット平均（リスクオン17セクターの等ウェイト）
   ・守りバスケット平均（リスクオフ4セクター）
   ・フロー最弱3セクター（シグナル日に資金流出が最も強い3つ）＝逆張り
   ・フロー最強3セクター（流入が最も強い3つ）＝順張り

 重要な前提（結果の読み方に影響）:
   ・シグナルは因果的（その日までのデータのみ）に検出されるため、
     ここでの検証はウォークフォワード検証と等価。未来のデータは使っていない。
   ・エントリーは「シグナル日の翌営業日の終値」（当日終値後に判明するため）
   ・売買コスト・税・配当は考慮しない
   ・シグナル同士の期間が重なることがある（各件は独立ではない）
============================================================
"""
from pathlib import Path
import sys

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_dashboard import compute, load_baskets, PRICES  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parent
INDICES = ROOT / "data" / "indices.parquet"
OUT_XLSX = Path(r"C:\Users\kawai\OneDrive\デスクトップ\株式投資\分析結果") / \
    "セクターローテーション_シグナルバックテスト.xlsx"

ENTRY_LAG = 1                      # シグナル日→エントリーまでの営業日数（翌日終値）
HORIZONS = {"1ヶ月": 20, "3ヶ月": 60, "6ヶ月": 120}   # 営業日

SIGNAL_LABELS = {
    "cross_up": "ゼロクロス↑（リスクオン転換）",
    "cross_dn": "ゼロクロス↓（リスクオフ転換）",
    "plus_break": "+σブレイク（過熱気味）",
    "minus_break": "−σブレイク（悲観極まる）",
    "accel": "加速アラート",
    "decel": "減速アラート",
}


def basket_logret() -> pd.DataFrame:
    """バスケット別の日次対数リターン（%）— build_dashboard と同じ計算"""
    baskets, labels, risk_off = load_baskets()
    px = pd.read_parquet(PRICES)
    close = px.pivot_table(index="Date", columns="Code", values="AdjC").sort_index()
    close = close.ffill(limit=3)
    ret = np.log(close).diff() * 100.0
    basket_ret = pd.DataFrame({
        name: ret[[c for c in codes if c in ret.columns]].mean(axis=1)
        for name, codes in baskets.items()
    })
    return basket_ret, labels, risk_off


def build_asset_logrets() -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """検証対象資産の日次対数リターン（%）と、セクター別リターンを返す"""
    basket_ret, labels, risk_off = basket_logret()
    risk_on = [c for c in basket_ret.columns if c not in risk_off]

    idx = pd.read_parquet(INDICES).set_index("Date").sort_index()
    idx_lr = np.log(idx).diff() * 100.0

    assets = pd.DataFrame({
        "日経平均": idx_lr["NIKKEI225"],
        "TOPIX": idx_lr["TOPIX"],
        "攻めバスケット平均": basket_ret[risk_on].mean(axis=1),
        "守りバスケット平均": basket_ret[risk_off].mean(axis=1),
    })
    assets = assets.reindex(basket_ret.index)
    return assets, basket_ret, labels


def fwd_return(logret: pd.Series, pos: int, h: int) -> float:
    """pos日目の終値で買い、h営業日後の終値で売った単純リターン（%）"""
    window = logret.iloc[pos + 1: pos + 1 + h]
    if len(window) < h or window.isna().any():
        return np.nan
    return (np.exp(window.sum() / 100.0) - 1.0) * 100.0


def main() -> None:
    frames, series, labels = compute()
    flow = frames["flow"]
    assets, basket_ret, _ = build_asset_logrets()
    dates = flow.index

    trades = []
    for sig_key, sig_label in SIGNAL_LABELS.items():
        sig_dates = series[sig_key][series[sig_key]].index
        for d in sig_dates:
            i = dates.get_loc(d)
            entry_i = i + ENTRY_LAG
            if entry_i >= len(dates):
                continue
            entry_d = dates[entry_i]

            # シグナル日のフローで最弱/最強3セクターを選ぶ
            f = flow.loc[d].sort_values()
            weakest3, strongest3 = list(f.index[:3]), list(f.index[-3:])

            targets = {
                "日経平均": assets["日経平均"],
                "TOPIX": assets["TOPIX"],
                "攻めバスケット平均": assets["攻めバスケット平均"],
                "守りバスケット平均": assets["守りバスケット平均"],
                "フロー最弱3セクター（逆張り）": basket_ret[weakest3].mean(axis=1),
                "フロー最強3セクター（順張り）": basket_ret[strongest3].mean(axis=1),
            }
            for asset_name, lr in targets.items():
                lr = lr.reindex(dates)
                pos = dates.get_loc(entry_d)
                row = {
                    "シグナル": sig_label,
                    "資産": asset_name,
                    "シグナル日": d.date(),
                    "エントリー日": entry_d.date(),
                }
                if "最弱" in asset_name:
                    row["選ばれたセクター"] = "、".join(labels[c] for c in weakest3)
                elif "最強" in asset_name:
                    row["選ばれたセクター"] = "、".join(labels[c] for c in strongest3)
                else:
                    row["選ばれたセクター"] = ""
                for h_label, h in HORIZONS.items():
                    row[h_label] = fwd_return(lr, pos, h)
                trades.append(row)

    df = pd.DataFrame(trades)

    # ベースライン: 全営業日で同じ保有をした場合の平均（シグナルに意味があるかの比較用）
    base = {}
    for asset_name in ["日経平均", "TOPIX", "攻めバスケット平均", "守りバスケット平均"]:
        lr = assets[asset_name].reindex(dates)
        for h_label, h in HORIZONS.items():
            vals = [fwd_return(lr, i, h) for i in range(len(dates))]
            vals = [v for v in vals if not np.isnan(v)]
            base[(asset_name, h_label)] = (np.mean(vals),
                                           np.mean([v > 0 for v in vals]) * 100)

    switch_rows, switch_trades = switching_strategy(series, assets, basket_ret, dates)
    write_excel(df, base, switch_rows, switch_trades)

    # コンソールにも要約を出す
    print(f"取引数: {len(df)}（シグナル{df['シグナル'].nunique()}種 × 資産6 × 発生回数）")
    for h_label in HORIZONS:
        pivot = df.pivot_table(values=h_label, index="シグナル", columns="資産", aggfunc="mean")
        print(f"\n===== {h_label}後の平均リターン（%）=====")
        print(pivot.round(2).to_string())


def switching_strategy(series, assets, basket_ret, dates):
    """加速アラートで買い、減速アラートで売るスイッチング戦略"""
    def exec_dates(key):
        out = []
        for d in series[key][series[key]].index:
            i = dates.get_loc(d) + ENTRY_LAG
            if i < len(dates):
                out.append(dates[i])
        return out

    events = sorted([(d, "B") for d in exec_dates("accel")] +
                    [(d, "S") for d in exec_dates("decel")])
    periods, pos, entry = [], 0, None
    for d, typ in events:
        if typ == "B" and pos == 0:
            pos, entry = 1, d
        elif typ == "S" and pos == 1 and d > entry:
            periods.append((entry, d, False))
            pos = 0
    if pos == 1:
        periods.append((entry, dates[-1], True))

    targets = {
        "日経平均": assets["日経平均"],
        "TOPIX": assets["TOPIX"],
        "攻めバスケット平均": assets["攻めバスケット平均"],
    }
    summary, trades = [], []
    for name, lr in targets.items():
        lr = lr.reindex(dates)
        total, wins, held = 1.0, 0, 0
        for e, x, open_ in periods:
            seg = lr.loc[e:x].iloc[1:]
            r = (np.exp(seg.sum() / 100.0) - 1.0) * 100
            total *= 1 + r / 100.0
            held += len(seg)
            if r > 0:
                wins += 1
            trades.append({"資産": name, "買い日": e.date(), "売り日": x.date(),
                           "保有営業日": len(seg), "リターン": r,
                           "状態": "保有中" if open_ else "決済済"})
        bh = (np.exp(lr.dropna().sum() / 100.0) - 1.0) * 100
        summary.append({"資産": name, "取引数": len(periods),
                        "勝率": wins / max(len(periods), 1) * 100,
                        "戦略の累積リターン": (total - 1) * 100,
                        "持ちっぱなし": bh,
                        "市場にいた割合": held / (len(dates) - 1) * 100})
    return summary, pd.DataFrame(trades)


def write_excel(df: pd.DataFrame, base: dict,
                switch_rows: list, switch_trades: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(name="Arial", bold=True)
    normal = Font(name="Arial")
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    note_font = Font(name="Arial", size=9, color="666666")

    # ---------- シート2: 取引一覧（先に作って行数を確定させる） ----------
    ws2 = wb.create_sheet("取引一覧")
    cols = ["シグナル", "資産", "シグナル日", "エントリー日", "選ばれたセクター",
            "1ヶ月", "3ヶ月", "6ヶ月"]
    for j, c in enumerate(cols, 1):
        cell = ws2.cell(1, j, c)
        cell.font = bold
        cell.fill = hdr_fill
    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, c in enumerate(cols, 1):
            v = row[c]
            if c in ("1ヶ月", "3ヶ月", "6ヶ月"):
                if pd.isna(v):
                    continue
                cell = ws2.cell(i, j, float(v) / 100.0)   # %は小数で保持
                cell.number_format = "0.0%"
            else:
                cell = ws2.cell(i, j, str(v))
            cell.font = normal
    n = len(df) + 1
    widths2 = [30, 26, 12, 12, 40, 9, 9, 9]
    for j, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ---------- シート1: サマリー ----------
    ws = wb.active
    ws.title = "サマリー"
    ws["A1"] = "セクターローテーション・シグナル バックテスト"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    period = f"{df['シグナル日'].min()} 〜 {df['シグナル日'].max()}"
    notes = [
        f"検証期間: {period}（シグナル発生日ベース・約5年）",
        "シグナルはその日までのデータのみで検出（未来のデータは不使用）。エントリーは翌営業日の終値",
        "リターン: エントリー終値から20/60/120営業日後（≒1/3/6ヶ月）の終値まで。コスト・税・配当は未考慮",
        "勝率: リターンがプラスだった割合。件数が少ないシグナルは偶然の影響が大きい点に注意",
        "シグナル同士の保有期間が重なることがあるため、各件は完全に独立した取引ではない",
        "サマリーの数値は「取引一覧」シートから数式で集計（取引一覧を編集すると自動で再計算される）",
    ]
    for k, t in enumerate(notes, 2):
        ws.cell(k, 1, "・" + t).font = note_font

    r0 = len(notes) + 3
    headers = ["シグナル", "資産", "件数",
               "1ヶ月平均", "1ヶ月勝率", "3ヶ月平均", "3ヶ月勝率", "6ヶ月平均", "6ヶ月勝率"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(r0, j, h)
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    sig_order = list(SIGNAL_LABELS.values())
    asset_order = ["日経平均", "TOPIX", "攻めバスケット平均", "守りバスケット平均",
                   "フロー最弱3セクター（逆張り）", "フロー最強3セクター（順張り）"]
    hcols = {"1ヶ月": "F", "3ヶ月": "G", "6ヶ月": "H"}

    r = r0 + 1
    for sig in sig_order:
        for asset in asset_order:
            ws.cell(r, 1, sig).font = normal
            ws.cell(r, 2, asset).font = normal
            crit = f'取引一覧!$A$2:$A${n},$A{r},取引一覧!$B$2:$B${n},$B{r}'
            ws.cell(r, 3, f'=COUNTIFS({crit})').font = normal
            for j, (h_label, col) in enumerate(hcols.items()):
                rng = f'取引一覧!${col}$2:${col}${n}'
                avg = ws.cell(r, 4 + j * 2,
                              f'=IFERROR(AVERAGEIFS({rng},{crit}),"")')
                avg.number_format = "+0.0%;-0.0%;0.0%"
                avg.font = normal
                win = ws.cell(r, 5 + j * 2,
                              f'=IFERROR(COUNTIFS({crit},{rng},">0")/'
                              f'COUNTIFS({crit},{rng},"<>"),"")')
                win.number_format = "0%"
                win.font = normal
            r += 1
        r += 1  # シグナル間に空行

    # ベースライン表
    r += 1
    ws.cell(r, 1, "【ベースライン】シグナルに関係なく毎日買った場合の平均（比較用・Pythonで全営業日から算出）").font = bold
    r += 1
    for j, h in enumerate(["資産", "", "",
                           "1ヶ月平均", "1ヶ月勝率", "3ヶ月平均", "3ヶ月勝率", "6ヶ月平均", "6ヶ月勝率"], 1):
        if h:
            cell = ws.cell(r, j, h)
            cell.font = bold
            cell.fill = hdr_fill
    r += 1
    for asset_name in ["日経平均", "TOPIX", "攻めバスケット平均", "守りバスケット平均"]:
        ws.cell(r, 1, asset_name).font = normal
        for j, h_label in enumerate(HORIZONS):
            mean_v, win_v = base[(asset_name, h_label)]
            c1 = ws.cell(r, 4 + j * 2, mean_v / 100.0)
            c1.number_format = "+0.0%;-0.0%;0.0%"
            c1.font = normal
            c2 = ws.cell(r, 5 + j * 2, win_v / 100.0)
            c2.number_format = "0%"
            c2.font = normal
        r += 1

    widths = [30, 26, 7, 11, 10, 11, 10, 11, 10]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(r0 + 1, 1)

    # ---------- シート3: スイッチング戦略 ----------
    ws3 = wb.create_sheet("スイッチング戦略")
    ws3["A1"] = "加速アラートで買い、減速アラートで売る戦略"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13)
    for k, t in enumerate([
        "加速アラートの翌営業日終値で買い、次の減速アラートの翌営業日終値で売る、を繰り返した場合",
        "保有中に出た追加の買いシグナルは無視。最後まで売りが出ていない場合は直近終値で評価",
        "売買コスト・税は未考慮（利確ごとに課税される口座では手取りはさらに減る）",
    ], 2):
        ws3.cell(k, 1, "・" + t).font = note_font

    r3 = 6
    sw_headers = ["資産", "取引数", "勝率", "戦略の累積リターン", "持ちっぱなし", "市場にいた割合"]
    for j, h in enumerate(sw_headers, 1):
        cell = ws3.cell(r3, j, h)
        cell.font = bold
        cell.fill = hdr_fill
    for i, row in enumerate(switch_rows, r3 + 1):
        ws3.cell(i, 1, row["資産"]).font = normal
        ws3.cell(i, 2, row["取引数"]).font = normal
        for j, key in enumerate(["勝率", "戦略の累積リターン", "持ちっぱなし", "市場にいた割合"], 3):
            c = ws3.cell(i, j, row[key] / 100.0)
            c.number_format = "0%" if key in ("勝率", "市場にいた割合") else "+0.0%;-0.0%;0.0%"
            c.font = normal

    r3 += len(switch_rows) + 3
    ws3.cell(r3, 1, "【取引明細】").font = bold
    r3 += 1
    tr_cols = ["資産", "買い日", "売り日", "保有営業日", "リターン", "状態"]
    for j, h in enumerate(tr_cols, 1):
        cell = ws3.cell(r3, j, h)
        cell.font = bold
        cell.fill = hdr_fill
    for i, (_, row) in enumerate(switch_trades.iterrows(), r3 + 1):
        for j, c in enumerate(tr_cols, 1):
            v = row[c]
            if c == "リターン":
                cell = ws3.cell(i, j, float(v) / 100.0)
                cell.number_format = "+0.0%;-0.0%;0.0%"
            elif c == "保有営業日":
                cell = ws3.cell(i, j, int(v))
            else:
                cell = ws3.cell(i, j, str(v))
            cell.font = normal
    for j, w in enumerate([20, 12, 12, 12, 11, 10], 1):
        ws3.column_dimensions[get_column_letter(j)].width = w

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Excel出力: {OUT_XLSX}")


if __name__ == "__main__":
    main()
